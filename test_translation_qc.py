"""
Headless tests for the Translation QC engine (translation_qc.py).

Uses ONLY synthetic data - client sheets must never enter the repo.
Covers the verdict golden set plus the precision & source-language rules.

Run:  python test_translation_qc.py
"""
import csv
import os
import sys
import tempfile
import importlib.util

_spec = importlib.util.spec_from_file_location("tqc", "translation_qc.py")
q = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(q)

_failures = []


def check(name, got, exp):
    if got != exp:
        _failures.append(name)
        print(f"[FAIL] {name}\n       got: {got}\n       exp: {exp}")
    else:
        print(f"[PASS] {name}")


LI = q.LanguageIdentifier(("en", "es"))


def v(src, tgt):
    return q.classify(src, tgt, LI, "en", "es")


def test_golden_synthetic():
    r = v("The market is rising quickly", "El mercado está subiendo rápidamente")
    check("golden.pass", (r.verdict, r.src_flag), ("PASS", False))

    r = v("La bolsa de valores de Corea del Sur", "La bolsa de valores de Corea del Sur")
    check("golden.burned_in", (r.verdict, r.severity, r.src_flag),
          ("SOURCE_ALREADY_TARGET", "NOTE", True))

    r = v("Los inversores están preocupados por la burbuja", "")
    check("golden.missing_with_note", (r.verdict, r.severity, r.src_flag),
          ("MISSING", "HIGH", True))

    r = v("The bubble is about to burst according to analysts",
          "The bubble is about to burst according to analysts")
    check("golden.untranslated", (r.verdict, r.severity), ("UNTRANSLATED_COPY", "HIGH"))


def test_precision_rule():
    # Shared-proper-noun rows must never be HIGH
    r = v("Youkyung Lee Reporter, Bloomberg News", "Youkyung Lee Reportera, Bloomberg News")
    check("precision.reportera_not_high", r.severity != "HIGH", True)
    r = v("Bloomberg ORIGINALS", "Bloomberg ORIGINALES")
    check("precision.originales_not_high", r.severity != "HIGH", True)


def test_cognate_false_positives():
    # Correctly-translated Spanish rows must not flag yellow just because the
    # Spanish looks like the English (cognates, single words, bylines/dates).
    r = v("Produced by", "Producido por")
    check("cognate.producido_por", (r.verdict, r.severity), ("PASS", "PASS"))
    r = v("Hosts", "Presentadores")
    check("cognate.presentadores", (r.verdict, r.severity), ("PASS", "PASS"))
    r = v("Bloomberg Buffalo Bills Add Toronto Sports Stars as Owners",
          "Vanessa Perdomo | 20 de dic. de 2024")
    check("cognate.byline_date", (r.verdict, r.severity), ("PASS", "PASS"))
    # But half-translated ("Spanglish") rows must STAY flagged:
    r = v("The market is rising", "El mercado is rising")
    check("cognate.spanglish_still_flagged", r.severity in ("MEDIUM", "HIGH"), True)
    # And an identical cognate word is still only VERIFY, never silent PASS:
    r = v("Editor", "Editor")
    check("cognate.identical_still_verify", (r.verdict, r.severity), ("VERIFY", "LOW"))


def test_verify_tier():
    # Identical brands surface as LOW VERIFY (Decision B), never silent, never HIGH
    for brand in ("SAMSUNG", "SK hynix", "Nasdaq", "OpenAI"):
        r = v(brand, brand)
        check(f"verify.{brand}", (r.verdict, r.severity), ("VERIFY", "LOW"))


def test_lang_mismatch():
    r = v("The stocks fell sharply on Monday morning",
          "The shares dropped hard when the market opened for the day")
    check("mismatch.english_in_target", (r.verdict, r.severity), ("LANG_MISMATCH", "MEDIUM"))


def test_non_translatable():
    check("nt.numbers", q._non_translatable("101%"), True)
    check("nt.allcaps", q._non_translatable("HYPERDRIVE"), True)
    check("nt.camelcase", q._non_translatable("OpenAI"), True)
    check("nt.prose", q._non_translatable("The market is rising"), False)


def test_end_to_end_and_nondestructive():
    rows = [
        ["speaker", "start_time", "end_time", "transcription", "translation"],
        ["On Screen", "0", "1", "Boom or bust?", "¿Auge o caída?"],
        ["On Screen", "0", "1", "SAMSUNG", "SAMSUNG"],
        ["On Screen", "0", "1", "The end is near for the bubble", ""],
    ]
    with tempfile.TemporaryDirectory() as tmp:
        pin = os.path.join(tmp, "sheet.csv")
        with open(pin, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)
        before = open(pin, "rb").read()
        res = q.run_qc(pin)
        after = open(pin, "rb").read()
        check("e2e.nondestructive", before == after, True)
        check("e2e.xlsx_written", os.path.exists(res["out"]), True)
        check("e2e.high_count", res["severity"].get("HIGH", 0), 1)
        check("e2e.rows", len(res["results"]), 3)
        from openpyxl import load_workbook
        wb = load_workbook(res["out"])
        check("e2e.sheets", wb.sheetnames, ["QC", "Summary"])
        hdr = [c.value for c in wb["QC"][1]]
        check("e2e.qc_columns", hdr[-6:],
              ["QC_Verdict", "QC_Severity", "QC_Confidence", "QC_TargetLang", "QC_SourceNote", "QC_Reason"])


def test_xlsx_input():
    # .xlsx sheets are accepted alongside .csv
    from openpyxl import Workbook, load_workbook
    with tempfile.TemporaryDirectory() as tmp:
        pin = os.path.join(tmp, "sheet.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "OST"
        ws.append(["speaker", "start_time", "end_time", "transcription", "translation"])
        ws.append(["On Screen", "0", "1", "Boom or bust?", "¿Auge o caída?"])
        ws.append(["On Screen", "0", "1", "SAMSUNG", "SAMSUNG"])
        ws.append(["On Screen", "0", "1", "The end is near for the bubble", ""])
        ws.append(["On Screen", 101, 2.5, "Scene 101", "Escena 101"])   # numeric cells -> text
        ws.append([None, None, None, None, None])                        # padding row -> dropped
        wb.save(pin)

        before = open(pin, "rb").read()
        res = q.run_qc(pin)
        after = open(pin, "rb").read()
        check("xlsx.nondestructive", before == after, True)
        check("xlsx.rows_padding_dropped", len(res["results"]), 4)
        check("xlsx.high_count", res["severity"].get("HIGH", 0), 1)
        check("xlsx.numeric_cell_str",
              res["data"]["rows"][3][1:3], ["101", "2.5"])
        check("xlsx.out_written", os.path.exists(res["out"]), True)
        out_wb = load_workbook(res["out"])
        check("xlsx.out_sheets", out_wb.sheetnames, ["QC", "Summary"])

    # header on a later sheet: first sheet unmappable, second sheet is the OST
    with tempfile.TemporaryDirectory() as tmp:
        pin = os.path.join(tmp, "multi.xlsx")
        wb = Workbook()
        wb.active.title = "Notes"
        wb.active.append(["just", "random", "columns"])
        ws2 = wb.create_sheet("OST")
        ws2.append(["transcription", "translation"])
        ws2.append(["Boom or bust?", "¿Auge o caída?"])
        wb.save(pin)
        res = q.run_qc(pin)
        check("xlsx.second_sheet_used", "OST" in res["data"]["encoding"], True)
        check("xlsx.second_sheet_rows", len(res["results"]), 1)


def test_refuse_overwriting_input():
    # run_qc must never write onto the input workbook (flag-only guarantee)
    from openpyxl import Workbook
    with tempfile.TemporaryDirectory() as tmp:
        pin = os.path.join(tmp, "sheet.xlsx")
        wb = Workbook()
        wb.active.append(["transcription", "translation"])
        wb.active.append(["Boom or bust?", "¿Auge o caída?"])
        wb.save(pin)
        try:
            q.run_qc(pin, out_path=pin)
            check("guard.same_path_refused", "no exception", "SystemExit")
        except SystemExit:
            check("guard.same_path_refused", "SystemExit", "SystemExit")


def test_reverse_direction():
    # LAS -> ENG ("vice versa"): Spanish source, English target expected
    li = q.LanguageIdentifier(("es", "en"))
    r = q.classify("El mercado está subiendo", "The market is rising", li, "es", "en")
    check("reverse.pass", r.verdict, "PASS")
    r = q.classify("El mercado está subiendo", "El mercado está subiendo", li, "es", "en")
    check("reverse.untranslated_high", (r.verdict, r.severity), ("UNTRANSLATED_COPY", "HIGH"))


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    test_golden_synthetic()
    test_precision_rule()
    test_cognate_false_positives()
    test_verify_tier()
    test_lang_mismatch()
    test_non_translatable()
    test_end_to_end_and_nondestructive()
    test_xlsx_input()
    test_refuse_overwriting_input()
    test_reverse_direction()
    print("\nRESULT:", "ALL PASS" if not _failures else f"{len(_failures)} FAILURE(S): {_failures}")
    sys.exit(1 if _failures else 0)
