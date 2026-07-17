"""
translation_qc.py - OST Translation QC (offline, flag-only).

Reads an OST sheet (.csv or .xlsx), flags rows whose target cell is not a real
translation (missing / untranslated copy / wrong language / verify), and writes
a colour-coded Excel copy. The source file is NEVER modified.

Implements the locked decisions (see PRD/PRD_2_AI_Implementation.md):
  A. Zero-dependency heuristic language ID (EN vs es-419); pluggable interface.
  B. Legitimately-identical cells (brands/numbers) -> low-severity "VERIFY".
  C. Output: highlighted Excel copy.
  D. Scale: up to ~50k rows (streamed read + memoized verdicts).
  E. Source-language notes (client feedback round 2): flags rows whose SOURCE
     cell already contains the target language (video burned-in text). The
     source cell and the adjacent translated cell are shown in a new colour
     (blue) so the pair is visually linked.
  F. Input formats (client feedback round 3): .xlsx workbooks are accepted
     alongside .csv - the first worksheet whose headers map both columns is
     used, fully-empty padding rows are dropped.

Usage:
  python translation_qc.py                       # auto-find a .csv/.xlsx in this folder
  python translation_qc.py "sheet.csv"
  python translation_qc.py "sheet.xlsx" --source en --target es --out report.xlsx
  python translation_qc.py "sheet.csv" --source-col transcription --target-col translation

Only `openpyxl` is required (for the Excel input/export); language detection is stdlib.
"""
import argparse
import csv
import difflib
import io
import os
import re
import sys
import unicodedata
from collections import Counter, namedtuple

# ======================================================================
# LANGUAGE PROFILES  (Decision A - stdlib heuristic; add languages here)
# ======================================================================
LANG_PROFILES = {
    "en": {
        "name": "English",
        "chars": set(),
        "stop": set((
            "the a an of and to in is are was were for with on by at as it this that from be or "
            "not you we they he she his her its their our your have has had will would can could "
            "should i but if then than so out up down over under new now"
        ).split()),
    },
    "es": {
        "name": "Spanish (es-419)",
        "chars": set("áéíóúñ¿¡üÁÉÍÓÚÑÜ"),
        "stop": set((
            "el la los las un una unos unas de del al que en y o a con por para más como se su sus "
            "lo le no sí pero está están son fue ser hay muy también sobre entre cuando desde sin ya "
            "este esta estos estas nos les nuestra nuestro acciones cierre"
        ).split()),
    },
}

_WORD_RE = re.compile(r"[^\W\d_]+", re.UNICODE)


def _tokens(s):
    return [t.lower() for t in _WORD_RE.findall(s or "")]


class LanguageIdentifier:
    """Tiny offline language guesser. Swap this class for a bundled model later."""

    def __init__(self, langs=("en", "es")):
        self.langs = [l for l in langs if l in LANG_PROFILES]

    def detect(self, text):
        """Return (lang_code, confidence 0..1). ('und', 0.0) when undecidable."""
        toks = _tokens(text)
        if not toks:
            return ("und", 0.0)
        scores = {}
        for l in self.langs:
            p = LANG_PROFILES[l]
            sw = sum(1 for t in toks if t in p["stop"])
            ch = sum(1 for c in (text or "") if c in p["chars"])
            scores[l] = sw + 2 * ch
        best = max(scores, key=scores.get)
        total = sum(scores.values())
        if total == 0 or scores[best] == 0:
            return ("und", 0.0)
        conf = scores[best] / (total + 1e-9)
        conf *= min(1.0, len(toks) / 4.0)   # short strings are less reliable
        return (best, round(conf, 3))


# ======================================================================
# NORMALISATION / SIMILARITY / TRANSLATABILITY
# ======================================================================
def _normalize(s):
    s = unicodedata.normalize("NFKC", s or "").lower().strip()
    s = re.sub(r"[^\w\s]", "", s, flags=re.UNICODE)
    return re.sub(r"\s+", " ", s)


def _similarity(a, b):
    na, nb = _normalize(a), _normalize(b)
    if not na and not nb:
        return 1.0
    return difflib.SequenceMatcher(None, na, nb).ratio()


# Starter allowlist of brands / tickers / proper nouns that legitimately stay
# identical across languages. In P1 this is a user-editable file (PRD Q-G).
BRAND_ALLOWLIST = set(x.lower() for x in [
    "samsung", "nasdaq", "kospi", "openai", "hyperdrive", "sk hynix",
    "bloomberg", "computex", "s&p 500", "youtube", "nvidia", "kospi intraday",
])


def _non_translatable(s):
    """True if the string is EXPECTED to stay identical across languages."""
    st = (s or "").strip()
    if not st:
        return False
    if st.lower() in BRAND_ALLOWLIST:
        return True
    if re.fullmatch(r"[\W\d\s_]+", st):                 # numbers / punctuation / symbols only
        return True
    letters = [c for c in st if c.isalpha()]
    if letters and all(c.isupper() for c in letters):   # ALLCAPS token(s), e.g. SAMSUNG
        return True
    if re.search(r"[a-z][A-Z]", st):                    # CamelCase brand, e.g. OpenAI
        return True
    if len(st) <= 3:                                    # too short to judge
        return True
    return False


# ======================================================================
# VERDICT ENGINE
# ======================================================================
Verdict = namedtuple(
    "Verdict",
    "verdict severity confidence lang reason src_lang src_flag",
    defaults=("und", False),   # src_lang, src_flag (source already in target language)
)


def classify(source, target, li, source_lang="en", target_lang="es",
             t_copy=0.90, t_partial=0.60, min_len=4, conf_floor=0.15):
    s = (source or "").strip()
    t = (target or "").strip()

    if not t and not s:
        return Verdict("EMPTY_ROW", "INFO", 1.0, "und", "Both cells empty")

    # --- Source-language note (client feedback): the video sometimes has
    # target-language text burned into the frame, so the SOURCE cell itself
    # is already in the target language. Detect it up front.
    src_lang_det, src_conf = li.detect(s)
    src_flag = bool(s) and src_lang_det == target_lang and src_conf >= conf_floor
    tgt_name = LANG_PROFILES.get(target_lang, {}).get("name", target_lang)

    if not t and s:
        v = Verdict("MISSING", "HIGH", 1.0, "und", "Target cell is empty")
        if src_flag:
            v = v._replace(src_lang=src_lang_det, src_flag=True,
                           reason=f"Target cell is empty | note: source is already {tgt_name} (burned-in text?) - copy may be intended")
        return v

    lang, conf = li.detect(t)
    sim = _similarity(s, t)
    nt = _non_translatable(t)

    # Source already in target language: an identical "translation" is EXPECTED
    # (the on-screen text was already Spanish), not lazy copying.
    if src_flag and sim >= t_copy:
        return Verdict("SOURCE_ALREADY_TARGET", "NOTE", src_conf, lang,
                       f"Source text is already {tgt_name} (burned-in on-screen text) - identical translation is expected",
                       src_lang_det, True)

    v = _classify_target(s, t, lang, conf, sim, nt,
                         source_lang, target_lang, tgt_name,
                         t_copy, t_partial, min_len, conf_floor)
    if src_flag:
        v = v._replace(src_lang=src_lang_det, src_flag=True,
                       reason=v.reason + f" | note: source already contains {tgt_name}")
    return v


def _classify_target(s, t, lang, conf, sim, nt, source_lang, target_lang,
                     tgt_name, t_copy, t_partial, min_len, conf_floor):
    """Target-side verdict logic (source-language note handled by classify)."""

    # Identical / near-identical to the source
    if sim >= t_copy:
        if nt:
            return Verdict("VERIFY", "LOW", conf, lang,
                           "Identical to source but looks like a brand/number/proper noun - confirm it should stay untranslated")
        # Only call it "untranslated" when the target is CONFIDENTLY still the source
        # language. High overlap alone is usually just shared proper nouns/tickers
        # (e.g. 'Youkyung Lee Reporter' -> 'Youkyung Lee Reportera').
        if lang == source_lang and conf >= conf_floor:
            src_name = LANG_PROFILES.get(source_lang, {}).get("name", source_lang)
            return Verdict("UNTRANSLATED_COPY", "HIGH", max(conf, 0.6), lang,
                           f"Target is ~identical to source ({sim:.0%}) and still reads as {src_name} - appears untranslated")
        return Verdict("VERIFY", "LOW", conf, lang,
                       f"High overlap with source ({sim:.0%}) but likely shared proper nouns - confirm the translatable part was translated")

    # Different text from here on
    if nt:
        return Verdict("VERIFY", "LOW", conf, lang, "Short / proper-noun content - confirm")

    if lang == source_lang and lang != target_lang and conf >= conf_floor:
        return Verdict("LANG_MISMATCH", "MEDIUM", conf, lang,
                       f"Target reads as {LANG_PROFILES.get(lang, {}).get('name', lang)}, expected {tgt_name}")

    if sim >= t_partial:
        return Verdict("PARTIAL_COPY", "LOW", conf, lang, f"Partial overlap with source ({sim:.0%})")

    if len(t) < min_len or conf < conf_floor:
        return Verdict("REVIEW", "LOW", conf, lang, "Too short / low-confidence to judge")

    if lang == target_lang:
        return Verdict("PASS", "PASS", conf, lang, f"Reads as {tgt_name}")

    return Verdict("REVIEW", "LOW", conf, lang, "Language unclear")


# ======================================================================
# SHEET INPUT  (.csv: encoding auto-detect + delimiter sniff | .xlsx via openpyxl)
# ======================================================================
ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
XLSX_EXTS = (".xlsx", ".xlsm", ".xltx", ".xltm")
SOURCE_HINTS = ["transcription", "source", "original", "text", "dialogue", "english", "eng", "en"]
TARGET_HINTS = ["translation", "target", "translated", "spanish", "espanol", "español", "es", "las", "es-419"]


def _read_text(path):
    raw = open(path, "rb").read()
    for enc in ENCODINGS:
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace"), "latin-1(replace)"


def _pick_column(headers, hints, explicit=None):
    if explicit:
        for i, h in enumerate(headers):
            if h.strip().lower() == explicit.strip().lower():
                return i
        raise SystemExit(f"Column '{explicit}' not found in {headers}")
    low = [h.strip().lower() for h in headers]
    for hint in hints:                       # exact match first
        if hint in low:
            return low.index(hint)
    for i, h in enumerate(low):              # then substring
        if any(hint in h for hint in hints):
            return i
    return None


def _cell_str(v):
    """Excel cell value -> text. Integral floats lose the fake '.0'."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def read_ost_xlsx(path, source_col=None, target_col=None):
    """
    Reads an .xlsx workbook: picks the first worksheet whose header row maps
    both source and target columns. Fully-empty padding rows (stale Excel
    dimensions) are dropped. Returns the same dict shape as the CSV reader,
    so everything downstream is format-agnostic.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        first = None
        for ws in wb.worksheets:
            raw = [[_cell_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
            raw = [r for r in raw if any(c.strip() for c in r)]
            if not raw:
                continue
            headers, rows = raw[0], raw[1:]
            if first is None:
                first = (ws.title, headers)
            try:
                si = _pick_column(headers, SOURCE_HINTS, source_col)
                ti = _pick_column(headers, TARGET_HINTS, target_col)
            except SystemExit:
                continue   # explicit column not on this sheet - try the next one
            if si is not None and ti is not None:
                return {"headers": headers, "rows": rows, "src_idx": si, "tgt_idx": ti,
                        "encoding": f"xlsx (sheet: {ws.title})", "delimiter": "n/a"}
        if first is None:
            raise SystemExit("Empty workbook.")
        raise SystemExit(f"Could not map source/target columns on any sheet.\n"
                         f"First sheet '{first[0]}' headers: {first[1]}\n"
                         f"Use --source-col / --target-col.")
    finally:
        wb.close()


def read_ost(path, source_col=None, target_col=None):
    """Reads an OST sheet (.csv or .xlsx) into a format-agnostic dict."""
    if os.path.splitext(path)[1].lower() in XLSX_EXTS:
        return read_ost_xlsx(path, source_col, target_col)
    text, enc = _read_text(path)
    try:
        delim = csv.Sniffer().sniff(text[:4096], delimiters=",;\t").delimiter
    except csv.Error:
        delim = ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    if not rows:
        raise SystemExit("Empty CSV.")
    headers = rows[0]
    si = _pick_column(headers, SOURCE_HINTS, source_col)
    ti = _pick_column(headers, TARGET_HINTS, target_col)
    if si is None or ti is None:
        raise SystemExit(f"Could not map source/target columns from headers: {headers}\n"
                         f"Use --source-col / --target-col.")
    return {"headers": headers, "rows": rows[1:], "src_idx": si, "tgt_idx": ti,
            "encoding": enc, "delimiter": delim}


# ======================================================================
# EXCEL OUTPUT  (Decision C - highlighted copy)
# ======================================================================
SEV_FILL = {
    "HIGH":   ("FFC7CE", "9C0006"),
    "MEDIUM": ("FFEB9C", "9C6500"),
    "LOW":    ("FFF2CC", "7F6000"),
    "PASS":   ("C6EFCE", "006100"),
    "INFO":   ("EDEDED", "3F3F3F"),
    "NOTE":   ("BDD7EE", "1F4E79"),   # blue: source already in target language (burned-in text)
}


def write_highlighted_xlsx(data, results, out_path, source_lang, target_lang):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    fills = {k: PatternFill("solid", fgColor=v[0]) for k, v in SEV_FILL.items()}
    fonts = {k: Font(color=v[1]) for k, v in SEV_FILL.items()}
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(color="FFFFFF", bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "QC"

    headers = list(data["headers"]) + ["QC_Verdict", "QC_Severity", "QC_Confidence", "QC_TargetLang", "QC_SourceNote", "QC_Reason"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")

    src_col = data["src_idx"] + 1        # 1-based column of the source (transcription) cell
    tgt_col = data["tgt_idx"] + 1        # 1-based column of the translation cell
    ncols = len(data["headers"])
    src_note_txt = "SOURCE IS ALREADY TARGET-LANG"
    for r, (row, v) in enumerate(zip(data["rows"], results), start=2):
        for c in range(ncols):
            ws.cell(row=r, column=c + 1, value=(row[c] if c < len(row) else ""))
        ws.cell(row=r, column=ncols + 1, value=v.verdict)
        ws.cell(row=r, column=ncols + 2, value=v.severity)
        ws.cell(row=r, column=ncols + 3, value=v.confidence)
        ws.cell(row=r, column=ncols + 4, value=v.lang)
        ws.cell(row=r, column=ncols + 5, value=(src_note_txt if v.src_flag else ""))
        ws.cell(row=r, column=ncols + 6, value=v.reason)

        style = v.severity if v.severity in SEV_FILL else "INFO"
        for col in (tgt_col, ncols + 1, ncols + 2):     # translation cell + verdict + severity
            ws.cell(row=r, column=col).fill = fills[style]
            ws.cell(row=r, column=col).font = fonts[style]

        # Client feedback: when the SOURCE itself is already in the target
        # language (burned-in on-screen text), paint the source cell AND the
        # adjacent translated cell in the linking blue so the pair stands out.
        if v.src_flag:
            for col in (src_col, tgt_col, ncols + 5):
                ws.cell(row=r, column=col).fill = fills["NOTE"]
                ws.cell(row=r, column=col).font = fonts["NOTE"]

    # Presentation
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data['rows']) + 1}"
    widths = {data["src_idx"] + 1: 46, data["tgt_idx"] + 1: 46, ncols + 5: 26, ncols + 6: 60}
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(i, 16)

    # Summary sheet
    s2 = wb.create_sheet("Summary")
    s2.append(["OST Translation QC - Summary"])
    s2.append([])
    s2.append(["Source language", source_lang])
    s2.append(["Target language", target_lang])
    s2.append(["Rows analysed", len(results)])
    s2.append(["Source already in target language", sum(1 for v in results if v.src_flag)])
    s2.append([])
    s2.append(["Verdict", "Count"])
    for verdict, n in Counter(v.verdict for v in results).most_common():
        s2.append([verdict, n])
    s2.append([])
    s2.append(["Severity", "Count"])
    for sev, n in Counter(v.severity for v in results).most_common():
        row_i = s2.max_row + 1
        s2.append([sev, n])
        if sev in SEV_FILL:
            s2.cell(row=row_i, column=1).fill = fills[sev]
    s2.column_dimensions["A"].width = 22
    s2.column_dimensions["B"].width = 12
    s2["A1"].font = Font(bold=True, size=14)

    wb.save(out_path)


# ======================================================================
# PROGRAMMATIC API (used by the SyncFlow Automator GUI)
# ======================================================================
def app_dir():
    """Folder of the running app: next to the EXE when frozen (PyInstaller),
    otherwise the folder of this source file. Used to find qc_allowlist.txt."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def load_allowlist_file(path):
    """Merge extra allowlist terms from a text file (one per line, # comments)."""
    try:
        if path and os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        BRAND_ALLOWLIST.add(line.lower())
    except Exception:
        pass


def run_qc(input_path, source_lang="en", target_lang="es",
           source_col=None, target_col=None, out_path=None, progress_cb=None):
    """
    Full QC pass: read -> classify (memoized) -> write highlighted xlsx.
    Returns {"out", "results", "data", "severity", "verdicts", "src_notes"}.
    progress_cb(done_rows, total_rows) is called every 500 rows if given.
    """
    data = read_ost(input_path, source_col, target_col)
    li = LanguageIdentifier((source_lang, target_lang))
    cache = {}
    results = []
    total = len(data["rows"])
    for i, row in enumerate(data["rows"]):
        s = row[data["src_idx"]] if data["src_idx"] < len(row) else ""
        t = row[data["tgt_idx"]] if data["tgt_idx"] < len(row) else ""
        key = (s, t)
        if key not in cache:
            cache[key] = classify(s, t, li, source_lang, target_lang)
        results.append(cache[key])
        if progress_cb and (i % 500 == 0 or i == total - 1):
            progress_cb(i + 1, total)
    out = out_path or (os.path.splitext(input_path)[0] + "_QC_highlighted.xlsx")
    # Flag-only guarantee (invariant I2): never write onto the input sheet.
    if os.path.abspath(out) == os.path.abspath(input_path):
        raise SystemExit("Refusing to overwrite the input sheet - choose a different output path.")
    write_highlighted_xlsx(data, results, out, source_lang, target_lang)
    return {
        "out": out,
        "results": results,
        "data": data,
        "severity": dict(Counter(v.severity for v in results)),
        "verdicts": dict(Counter(v.verdict for v in results)),
        "src_notes": sum(1 for v in results if v.src_flag),
    }


# ======================================================================
# MAIN
# ======================================================================
def main(argv=None):
    # Never let console encoding crash the tool (Windows consoles are often cp1252).
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    ap = argparse.ArgumentParser(description="OST Translation QC (offline, flag-only).")
    ap.add_argument("input", nargs="?", help="OST sheet path (.csv or .xlsx; defaults to the first one in this folder)")
    ap.add_argument("--source", default="en", help="source language code (default en)")
    ap.add_argument("--target", default="es", help="target language code (default es)")
    ap.add_argument("--source-col", help="explicit source column header")
    ap.add_argument("--target-col", help="explicit target column header")
    ap.add_argument("--out", help="output .xlsx path")
    args = ap.parse_args(argv)

    inp = args.input
    if not inp:
        here = os.path.dirname(os.path.abspath(__file__))
        exts = (".csv",) + XLSX_EXTS
        sheets = sorted(
            (f for f in os.listdir(here)
             if f.lower().endswith(exts) and not f.lower().endswith("_qc_highlighted.xlsx")),
            key=lambda f: (not f.lower().endswith(".csv"), f.lower()))   # prefer .csv, then A-Z
        if not sheets:
            raise SystemExit("No sheet given and no .csv/.xlsx found in this folder.")
        inp = os.path.join(here, sheets[0])
        print(f"[i] Auto-selected: {os.path.basename(inp)}")

    load_allowlist_file(os.path.join(app_dir(), "qc_allowlist.txt"))

    res = run_qc(inp, args.source, args.target, args.source_col, args.target_col, args.out)
    data = res["data"]
    print(f"[i] Encoding: {data['encoding']}  delimiter: {data['delimiter']!r}")
    print(f"[i] Source col: '{data['headers'][data['src_idx']]}'  ->  Target col: '{data['headers'][data['tgt_idx']]}'")
    print(f"[i] Rows: {len(data['rows'])}")

    print("\n=== QC SUMMARY ===")
    print("Verdicts :", res["verdicts"])
    print("Severity :", res["severity"])
    flagged = sum(n for k, n in res["severity"].items() if k in ("HIGH", "MEDIUM"))
    print(f"Action-needed (High+Medium): {flagged} row(s)")
    print(f"Source-already-target notes (blue): {res['src_notes']} row(s)")
    print(f"[OK] Highlighted Excel written: {res['out']}")
    print("[OK] Source file untouched (flag-only).")


if __name__ == "__main__":
    main()
